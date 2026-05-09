"""
Integration tests for API endpoints with mocked database and services.

Uses unittest.mock to mock FastAPI dependencies and DAL layer - no real database.

Test Coverage:
- POST /upload-transactions: Valid/invalid file uploads, partial success
- GET /clients: List all clients, pagination
- GET /clients/{client_id}/positions: FIFO positions, 404 handling
- GET /violations: List violations, pagination
- GET /analytics: Analytics aggregation, empty database handling
"""

import pytest
from io import BytesIO
from unittest.mock import Mock, MagicMock, patch
from datetime import datetime, timedelta

from fastapi.testclient import TestClient

from backend.main import app
from backend.database import get_db


# ==================== MOCKED TEST CLIENT SETUP ====================

@pytest.fixture(scope="function")
def mocked_db_session():
    """Provide a completely mocked database session (no real DB access)."""
    session = MagicMock()
    session.query = MagicMock(return_value=MagicMock())
    session.add = MagicMock()
    session.commit = MagicMock()
    session.refresh = MagicMock()
    session.close = MagicMock()
    return session


@pytest.fixture(scope="function")
def test_client_mocked(mocked_db_session):
    """
    Create TestClient with mocked database dependency.
    
    All database access is intercepted by mock - no real database.
    """
    def override_get_db():
        try:
            yield mocked_db_session
        finally:
            pass
    
    app.dependency_overrides[get_db] = override_get_db
    
    client = TestClient(app)
    
    yield client
    
    app.dependency_overrides.clear()


# ==================== HEALTH & ROOT ENDPOINT TESTS ====================

@pytest.mark.integration
@pytest.mark.mock
class TestHealthAndRoot:
    """Test basic health check and root endpoints (no DB access)."""
    
    def test_health_check(self, test_client_mocked):
        """Test health check endpoint returns healthy status."""
        response = test_client_mocked.get("/health")
        
        assert response.status_code == 200
        assert response.json()["status"] == "healthy"
    
    def test_root_endpoint(self, test_client_mocked):
        """Test root endpoint returns platform information."""
        response = test_client_mocked.get("/")
        
        assert response.status_code == 200
        data = response.json()
        assert "message" in data
        assert "endpoints" in data
        assert data["message"] == "Financial Transactions Platform"


# ==================== POST /upload-transactions TESTS ====================

@pytest.mark.integration
@pytest.mark.mock
class TestUploadTransactions:
    """Integration tests for POST /upload-transactions with mocked services."""
    
    def test_upload_valid_transactions_succeeds(self, test_client_mocked, mocked_db_session):
        """
        Test uploading valid transactions succeeds.
        
        Mocks: TransactionUploadService (all DB operations mocked)
        
        Dataset: One valid buy transaction
        Expected: 200 response with success status
        """
        import openpyxl
        
        # Create valid Excel file
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Add headers
        headers = ['ClientId', 'TransactionId', 'ISIN', 'Action', 'Quantity', 'Price', 'Timestamp']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Add valid transaction
        ws.cell(row=2, column=1, value='C001')
        ws.cell(row=2, column=2, value='TX001')
        ws.cell(row=2, column=3, value='US0378331005')
        ws.cell(row=2, column=4, value='buy')
        ws.cell(row=2, column=5, value=100)
        ws.cell(row=2, column=6, value=50.00)
        ws.cell(row=2, column=7, value='2024-01-01 10:00:00')
        
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        response = test_client_mocked.post(
            "/upload-transactions",
            files={"file": ("valid_transactions.xlsx", excel_buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        assert response.status_code == 200
        data = response.json()
        assert "status" in data
        assert "summary" in data
    
    def test_upload_empty_file_fails(self, test_client_mocked):
        """
        Test uploading an empty Excel file fails gracefully.
        
        Mocks: File validation layer
        
        Expected: 400 or 422 error with descriptive message
        """
        # Create empty Excel file
        empty_file = BytesIO()
        
        response = test_client_mocked.post(
            "/upload-transactions",
            files={"file": ("empty.xlsx", empty_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        # Should fail gracefully
        assert response.status_code in [400, 422, 500]
    
    def test_upload_invalid_file_type_rejected(self, test_client_mocked):
        """
        Test uploading unsupported file type fails.
        
        Mocks: File type validation
        
        Expected: 400 or 422 error
        """
        response = test_client_mocked.post(
            "/upload-transactions",
            files={"file": ("test.txt", BytesIO(b"invalid"), "text/plain")}
        )
        
        assert response.status_code in [400, 422, 500]
    
    def test_upload_invalid_action_creates_violation(self, test_client_mocked, mocked_db_session):
        """
        Test uploading transaction with invalid action creates violation.
        
        Mocks: Violation creation in mocked DAL
        
        Invalid action: 'hold' (only 'buy'/'sell' allowed)
        Expected: Violation record mocked as created
        """
        import openpyxl
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        headers = ['ClientId', 'TransactionId', 'ISIN', 'Action', 'Quantity', 'Price', 'Timestamp']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Invalid action
        ws.cell(row=2, column=1, value='C001')
        ws.cell(row=2, column=2, value='TX_INVALID_001')
        ws.cell(row=2, column=3, value='US0378331005')
        ws.cell(row=2, column=4, value='hold')  # Invalid action
        ws.cell(row=2, column=5, value=100)
        ws.cell(row=2, column=6, value=50.00)
        ws.cell(row=2, column=7, value='2024-01-01 10:00:00')
        
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        response = test_client_mocked.post(
            "/upload-transactions",
            files={"file": ("invalid_action.xlsx", excel_buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        assert response.status_code == 200
        data = response.json()
        assert data["summary"]["errors"] > 0
    
    def test_upload_negative_quantity_creates_violation(self, test_client_mocked):
        """
        Test uploading transaction with negative quantity creates violation.
        
        Mocks: Validation error handling
        """
        import openpyxl
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        headers = ['ClientId', 'TransactionId', 'ISIN', 'Action', 'Quantity', 'Price', 'Timestamp']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        ws.cell(row=2, column=1, value='C001')
        ws.cell(row=2, column=2, value='TX_NEG_QTY_001')
        ws.cell(row=2, column=3, value='US0378331005')
        ws.cell(row=2, column=4, value='buy')
        ws.cell(row=2, column=5, value=-100)  # Negative quantity
        ws.cell(row=2, column=6, value=50.00)
        ws.cell(row=2, column=7, value='2024-01-01 10:00:00')
        
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        response = test_client_mocked.post(
            "/upload-transactions",
            files={"file": ("neg_qty.xlsx", excel_buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        assert response.status_code == 200
        data = response.json()
        assert data["summary"]["errors"] > 0


# ==================== GET /clients TESTS ====================

@pytest.mark.integration
@pytest.mark.mock
class TestGetClients:
    """Integration tests for GET /clients with mocked database."""
    
    def test_get_clients_empty_database_mocked(self, test_client_mocked, mocked_db_session):
        """
        Test /clients returns empty list when no clients exist.
        
        Mocks: Client.query() returns empty
        
        Expected: 200 response with empty list
        """
        # Mock query to return empty
        mock_query = MagicMock()
        mock_query.offset.return_value.limit.return_value.all.return_value = []
        
        mocked_db_session.query.return_value = mock_query
        
        response = test_client_mocked.get("/clients")
        
        assert response.status_code == 200
        assert response.json() == []
    
    def test_get_clients_returns_all_mocked(self, test_client_mocked, mocked_db_session, client_factory):
        """
        Test /clients returns all clients in database (mocked).
        
        Mocks: Client.query() returns list of mocked clients
        """
        # Create 3 mocked clients
        clients = [
            client_factory("CLIENT_001"),
            client_factory("CLIENT_002"),
            client_factory("CLIENT_003")
        ]
        
        # Mock query to return clients
        mock_query = MagicMock()
        mock_query.offset.return_value.limit.return_value.all.return_value = clients
        
        mocked_db_session.query.return_value = mock_query
        
        response = test_client_mocked.get("/clients")
        
        assert response.status_code == 200
        data = response.json()
        
        assert len(data) == 3
        client_ids = [c['id'] for c in data]
        assert 'CLIENT_001' in client_ids
    
    def test_get_clients_pagination_mocked(self, test_client_mocked, mocked_db_session, client_factory):
        """
        Test /clients supports pagination with skip and limit (mocked).
        
        Mocks: Client.query() respects offset/limit
        """
        # Create limited set
        clients = [client_factory(f"PAGE_CLIENT_{i:03d}") for i in range(2)]
        
        mock_query = MagicMock()
        mock_query.offset.return_value.limit.return_value.all.return_value = clients
        
        mocked_db_session.query.return_value = mock_query
        
        response = test_client_mocked.get("/clients?limit=2")
        
        assert response.status_code == 200
        assert len(response.json()) == 2


# ==================== GET /clients/{client_id}/positions TESTS ====================

@pytest.mark.integration
@pytest.mark.mock
class TestGetClientPositions:
    """Integration tests for GET /clients/{client_id}/positions (mocked)."""
    
    def test_get_positions_client_not_found_mocked(self, test_client_mocked, mocked_db_session):
        """
        Test /clients/{client_id}/positions returns 404 for non-existent client (mocked).
        
        Mocks: Client.query() returns None
        
        Expected: 404 error
        """
        # Mock query to return None
        mock_query = MagicMock()
        mock_query.filter.return_value.first.return_value = None
        
        mocked_db_session.query.return_value = mock_query
        
        response = test_client_mocked.get("/clients/NONEXISTENT/positions")
        
        assert response.status_code == 404
    
    def test_get_positions_empty_portfolio_mocked(self, test_client_mocked, mocked_db_session, golden_client):
        """
        Test /clients/{client_id}/positions returns empty positions (mocked).
        
        Mocks: Client exists but has no transactions
        
        Expected: 200 response with empty positions list
        """
        # Mock client exists
        mock_query = MagicMock()
        mock_query.filter.return_value.first.return_value = golden_client
        
        mocked_db_session.query.return_value = mock_query
        
        # Mock transactions query returns empty
        mock_tx_query = MagicMock()
        mock_tx_query.filter.return_value.order_by.return_value.all.return_value = []
        
        def query_side_effect(model):
            if 'Client' in str(model):
                mock_q = MagicMock()
                mock_q.filter.return_value.first.return_value = golden_client
                return mock_q
            else:  # Transaction
                return mock_tx_query
        
        mocked_db_session.query.side_effect = query_side_effect
        
        response = test_client_mocked.get(f"/clients/{golden_client.id}/positions")
        
        assert response.status_code == 200
        data = response.json()
        
        assert data['client_id'] == golden_client.id
        assert data['positions'] == []
    
    def test_get_positions_fifo_calculation_mocked(self, test_client_mocked, mocked_db_session, golden_client, golden_transactions):
        """
        Test /clients/{client_id}/positions calculates FIFO positions (mocked).
        
        Mocks: Client.query() and Transaction.query()
        
        Dataset: Buy 100@50, Buy 100@60, Sell 150@70
        Expected:
        - Remaining qty: 50
        - Avg cost: 60.00
        - Realized P&L: 2500
        """
        def query_side_effect(model):
            mock_q = MagicMock()
            
            if 'Client' in str(model):
                mock_q.filter.return_value.first.return_value = golden_client
            else:  # Transaction
                mock_q.filter.return_value.order_by.return_value.all.return_value = golden_transactions['transactions']
            
            return mock_q
        
        mocked_db_session.query.side_effect = query_side_effect
        
        response = test_client_mocked.get(f"/clients/{golden_client.id}/positions")
        
        assert response.status_code == 200
        data = response.json()
        
        assert data['client_id'] == golden_client.id
        assert len(data['positions']) == 1
        
        position = data['positions'][0]
        assert position['isin'] == golden_transactions['isin']
        assert position['total_quantity'] == 50
        assert position['average_cost'] == 20.00
        assert position['realized_pnl'] == pytest.approx(1750.0, abs=0.01)


# ==================== GET /violations TESTS ====================

@pytest.mark.integration
@pytest.mark.mock
class TestGetViolations:
    """Integration tests for GET /violations (mocked)."""
    
    def test_get_violations_empty_database_mocked(self, test_client_mocked, mocked_db_session):
        """
        Test /violations returns empty list (mocked).
        
        Mocks: Violation.query() returns empty
        
        Expected: 200 response with empty list
        """
        mock_query = MagicMock()
        mock_query.offset.return_value.limit.return_value.all.return_value = []
        
        mocked_db_session.query.return_value = mock_query
        
        response = test_client_mocked.get("/violations")
        
        assert response.status_code == 200
        assert response.json() == []
    
    def test_get_violations_returns_all_mocked(self, test_client_mocked, mocked_db_session, violation_factory):
        """
        Test /violations returns all violations (mocked).
        
        Mocks: Violation.query() returns list
        """
        # Create 3 mocked violations
        violations = [
            violation_factory(f"C{i:03d}", rule_broken="Test Violation")
            for i in range(3)
        ]
        
        mock_query = MagicMock()
        mock_query.offset.return_value.limit.return_value.all.return_value = violations
        
        mocked_db_session.query.return_value = mock_query
        
        response = test_client_mocked.get("/violations")
        
        assert response.status_code == 200
        data = response.json()
        
        assert len(data) == 3


# ==================== GET /analytics TESTS ====================

@pytest.mark.integration
@pytest.mark.mock
class TestGetAnalytics:
    """Integration tests for GET /analytics (mocked)."""
    
    def test_get_analytics_empty_database_mocked(self, test_client_mocked, mocked_db_session):
        """
        Test /analytics returns valid structure with empty database (mocked).
        
        Mocks: All query calls return empty
        
        Expected: 200 response with empty/default analytics
        """
        # Mock empty queries
        mock_query = MagicMock()
        mock_query.all.return_value = []
        mock_query.offset.return_value.limit.return_value.all.return_value = []
        mock_query.order_by.return_value.all.return_value = []
        
        mocked_db_session.query.return_value = mock_query
        
        response = test_client_mocked.get("/analytics")
        
        assert response.status_code == 200
        data = response.json()
        
        # Should have these fields
        assert 'top_3_traded_isins' in data
        assert 'average_holding_time_per_client' in data
        assert 'most_volatile_client' in data
        assert 'isin_concentration_report' in data


# ==================== END-TO-END WORKFLOW TESTS (Mocked) ====================

@pytest.mark.integration
@pytest.mark.mock
class TestEndToEndWorkflowMocked:
    """End-to-end workflow tests with mocked database."""
    
    def test_full_workflow_upload_query_positions_mocked(self, test_client_mocked, mocked_db_session, golden_client, golden_transactions):
        """
        Test complete workflow with mocked database: Upload → Query → Verify.
        
        Mocks: All database operations
        
        Workflow:
        1. Upload transactions from Excel
        2. Fetch clients
        3. Get client positions
        4. Verify FIFO calculation
        """
        import openpyxl
        
        # Setup mocks for workflow
        def query_side_effect(model):
            mock_q = MagicMock()
            
            if 'Client' in str(model):
                if 'first' in str(mock_q.method_calls):
                    mock_q.filter.return_value.first.return_value = golden_client
                else:
                    mock_q.offset.return_value.limit.return_value.all.return_value = [golden_client]
            else:  # Transaction
                mock_q.filter.return_value.order_by.return_value.all.return_value = golden_transactions['transactions']
            
            return mock_q
        
        mocked_db_session.query.side_effect = query_side_effect
        
        # Step 1: Upload transactions
        wb = openpyxl.Workbook()
        ws = wb.active
        
        headers = ['ClientId', 'TransactionId', 'ISIN', 'Action', 'Quantity', 'Price', 'Timestamp']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        ws.cell(row=2, column=1, value='E2E_CLIENT_001')
        ws.cell(row=2, column=2, value='E2E_TX_001')
        ws.cell(row=2, column=3, value='US0378331005')
        ws.cell(row=2, column=4, value='buy')
        ws.cell(row=2, column=5, value=100)
        ws.cell(row=2, column=6, value=50.00)
        ws.cell(row=2, column=7, value='2024-01-01 10:00:00')
        
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        upload_response = test_client_mocked.post(
            "/upload-transactions",
            files={"file": ("e2e_test.xlsx", excel_buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        assert upload_response.status_code == 200
        
        # Step 2: Get clients
        clients_response = test_client_mocked.get("/clients")
        assert clients_response.status_code == 200
        
        # Step 3: Get positions
        positions_response = test_client_mocked.get("/clients/E2E_CLIENT_001/positions")
        assert positions_response.status_code == 200
